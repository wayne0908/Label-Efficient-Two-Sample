import numpy as np 
import os
import pdb
import matplotlib.pyplot as plt 
import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
from sklearn.cluster import KMeans
from statsmodels.stats.proportion import proportion_confint
from sklearn import metrics
from random import randint
import sys
sys.path.append(os.getcwd() + "/Statsticscollection/")
from ComputeStatistics import *

def SetPltProp(ax, xn = None, yn=None, title=None, grid = True, bbox_to_anchor=None, legend = True, pos = 'upper left', borderpad=None):
	fontsize = 14
	for axis in ['top','bottom','left','right']:
		ax.spines[axis].set_linewidth(2.5)

	for tick in ax.xaxis.get_major_ticks():
	    tick.label1.set_fontsize(fontsize)
	    tick.label1.set_fontweight('bold')
	for tick in ax.yaxis.get_major_ticks():
	    tick.label1.set_fontsize(fontsize)
	    tick.label1.set_fontweight('bold') 

	if legend:  
		if bbox_to_anchor == None: 
			ax.legend(loc=pos, shadow=True, prop={'weight':'bold', 'size':10},  borderpad=borderpad)
		else:
			ax.legend(loc=pos, shadow=True, bbox_to_anchor =bbox_to_anchor, prop={'weight':'bold', 'size':10}, borderpad=borderpad)
	if grid:
		ax.grid(linewidth='1.5', linestyle='dashed')
	if xn != None:
		ax.set_xlabel(xn, fontweight='bold')
	if yn != None:
		ax.set_ylabel(yn, fontweight='bold')
	if title != None:
		ax.set_title(title, fontweight='bold')
	return ax

def SetScatterProp(ax, xn=None, yn=None, title=None, legend = True, Loc = 'upper right'):
	fontsize = 14
	for axis in ['top','bottom','left','right']:
		ax.spines[axis].set_linewidth(2.5)

	ax.set_yticklabels([])
	ax.set_xticklabels([])

	if legend:   
		ax.legend(loc=Loc, shadow=True, prop={'weight':'bold'})

	if title != None:
		ax.set_title(title, fontweight='bold')
	return ax

def DrawData(Args, data, BER):
	"""
	Args: parser(). Parameter options
	data: array. Dataset
	BER: Bayes error rate
	"""
	print('Drawing synthetic dataset...')
	Path = os.getcwd() + '/Figures/%s/Data/'%Args.DataType
	if not os.path.exists(Path):
		os.makedirs(Path)
	Fig = plt.figure(); ax = Fig.gca()

	Feat = data[:, :-1]
	Label = data[:, -1]
		
	ax.scatter(Feat[Label == 0, 0], Feat[Label == 0, 1], c = 'r', label = 'Class 0', s = 30)
	ax.scatter(Feat[Label == 1, 0], Feat[Label == 1, 1], c = 'b', label = 'Class 1', s = 30)
	plt.axis(aspect='equal')
	ax = SetScatterProp(ax, legend = True, Loc = 'upper left')
	ax.set_aspect(aspect='equal')
	Fig.savefig(Path + 'SynSep%.2fDel%.2fSize%d.png'%(Args.Sep, Args.Del, Args.S), bbox_inches='tight')
	plt.close('all')

def GetDirectory(Args):

	if Args.DataType == 'Syn':
		StatsPath1 = os.getcwd() + '/Stats/%s/D%dSep%.2fDel%.2fSize%d/%s/Alpha%.2f/'%(Args.DataType, Args.FeatLen, Args.Sep, Args.Del, Args.S, Args.TestType, Args.alpha)
		FigurePath1 = os.getcwd() + '/Figures/%s/D%dSep%.2fDel%.2fSize%d/%s/Alpha%.2f/'%(Args.DataType, Args.FeatLen, Args.Sep, Args.Del, Args.S, Args.TestType, Args.alpha)
	else:
		StatsPath1 = os.getcwd() + '/Stats/%s/%s/'%(Args.DataType, Args.TestType)
		FigurePath1 = os.getcwd() + '/Figures/%s/%s/'%(Args.DataType, Args.TestType)
	if Args.TestType == 'Chen':
		StatsPath1+='(M=%d)'%(Args.ChenM)
		FigurePath1+='(M=%d)'%(Args.ChenM)

	"""
	Load query index
	"""
	if Args.DataType == 'Syn':
		QueryLoadPath = os.getcwd() + '/Stats/%s/D%dSep%.2fDel%.2fSize%d/'%(Args.DataType, Args.FeatLen, Args.Sep, Args.Del, Args.S)
		QueryLoadFigPath = os.getcwd() + '/Figures/%s/D%dSep%.2fDel%.2fSize%d/'%(Args.DataType, Args.FeatLen, Args.Sep, Args.Del, Args.S)
	else:
		QueryLoadPath = os.getcwd() + '/Stats/%s/'%(Args.DataType)
		QueryLoadFigPath = os.getcwd() + '/Figures/%s/'%(Args.DataType)
	return StatsPath1, FigurePath1, QueryLoadPath, QueryLoadFigPath


def PlotStatsDistri2(args, StatsPath, FigurePath, StatsName):
	"""
	For permutation distribution generation
	Args: input argument 
	StatsPath: str. Path to the stats directory  
	FigurePath: str. Path to the figure directory
	StatsName: str. Name of stats
	"""
	stats=np.load(StatsPath + 'CutEdgeNumTrial%d.npy'%(args.Trial))
	
	PVal = stats[0]; Reject=stats[1]

	np.save(StatsPath + 'Reject%d.npy'%args.Trial, Reject);np.save(StatsPath + 'PVal%d.npy'%args.Trial, PVal)
	WriteToExcel(StatsPath, StatsName, Reject, PVal, args.Trial)


def PlotFDivergence(args, StatsPath, FigurePath):
	Per = np.arange(args.Interval, args.Per + args.Interval, args.Interval); 
	FigurePath2 = FigurePath + 'MI/'; 
	if not os.path.exists(FigurePath2):
		os.makedirs(FigurePath2)
	PassiveMI = np.zeros(len(Per)); EUMI = np.zeros(len(Per)); EUMI2 = np.zeros(len(Per)); USMI = np.zeros(len(Per))
	PassiveMIstd = np.zeros(len(Per)); EUMIstd = np.zeros(len(Per)); EUMI2std = np.zeros(len(Per)); USMIstd = np.zeros(len(Per))
	PassiveMIV = np.zeros(args.Trial); EUMIV = np.zeros(args.Trial); EUMI2V = np.zeros(args.Trial); USMIV = np.zeros(args.Trial)

	Path1 = StatsPath + 'QueryIndex(ClsUncertaintyQuery)/%s/MI/'%('Passive')
	Path2 = StatsPath + 'QueryIndex(ClsUncertaintyQuery)/%s/%s/MI/InitTrainingSize%d/'%('Uncertainty', args.cls, args.InitSize)
	Path3 = StatsPath + 'QueryIndex(ClsUncertaintyQuery)/%s/%s/MI/InitTrainingSize%d/'%('Certainty', args.cls, args.InitSize)
	Path4 = StatsPath + 'QueryIndex(ClsUncertaintyQuery)/%s/%s/MI/InitTrainingSize%d/'%('Bimodal', args.cls, args.InitSize)
	pwb=load_workbook(Path1 + 'MI.xlsx', data_only=True);pws=pwb["FDivergence"];
	ewb=load_workbook(Path2 + 'MI.xlsx', data_only=True);ews=ewb["FDivergence"];
	uwb=load_workbook(Path3 + 'MI.xlsx', data_only=True);uws=uwb["FDivergence"];
	ewb2=load_workbook(Path4 + 'MI.xlsx', data_only=True);ews2=ewb2["FDivergence"];
	Alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

	for i in range(len(Per)):
		for t in range(args.Trial):
			PassiveMI[i]+= pws['%s%d'%(Alphabet[i], t + 1)].value; PassiveMIV[t] = pws['%s%d'%(Alphabet[i], t + 1)].value
			EUMI[i]+=ews['%s%d'%(Alphabet[i], t + 1)].value; EUMIV[t] = ews['%s%d'%(Alphabet[i], t + 1)].value
			USMI[i]+=uws['%s%d'%(Alphabet[i], t + 1)].value; USMIV[t] = uws['%s%d'%(Alphabet[i], t + 1)].value
			EUMI2[i]+=ews2['%s%d'%(Alphabet[i], t + 1)].value; EUMI2V[t] = ews2['%s%d'%(Alphabet[i], t + 1)].value
		PassiveMIstd[i] = np.std(PassiveMIV); EUMIstd[i]=np.std(EUMIV);
		EUMI2std[i] =np.std(EUMI2V); USMIstd[i] = np.std(USMIV)

	PassiveMI /=args.Trial; EUMI /=args.Trial;USMI /=args.Trial;EUMI2 /=args.Trial
	
	Fig = plt.figure(); ax = Fig.gca()
	
	ax.plot(Per, PassiveMI, marker = 'o', c='g', markersize = 8, label='Passive', linewidth = 3); 
	ax.fill_between(Per, PassiveMI + PassiveMIstd, PassiveMI - PassiveMIstd, color='g', alpha = 0.15); 
	ax.plot(Per, USMI, marker = 'o', c='r',markersize = 8, label='Certainty', linewidth = 3); 
	ax.fill_between(Per, USMI + USMIstd, USMI - USMIstd, color='r', alpha = 0.15); 
	ax.plot(Per, EUMI, marker = 'o', c='b', markersize = 8, label='Uncertainty', linewidth = 3); 
	ax.fill_between(Per, EUMI + EUMIstd, EUMI - EUMIstd, color='b', alpha = 0.15); 
	ax.plot(Per, EUMI2, marker = 'o', c='c', markersize = 8, label='Bimodal(Proposed)', linewidth = 3); 
	ax.fill_between(Per, EUMI2 + EUMI2std, EUMI2 - EUMI2std, color='c', alpha = 0.15); 
	ax = SetPltProp(ax, xn='Proportion of queries ', yn='f-divergence', pos='upper right')
	Fig.savefig(FigurePath2 + 'FD%s.png'%args.cls, bbox_inches='tight')
	plt.close('all')
		
def PlotResultsTrend(args, StatsPath, FigurePath):
	Per = np.arange(args.Interval, args.Per + args.Interval, args.Interval); 
	# Per = np.arange(args.Interval, args.Per, args.Interval); 
	FigurePath2 = FigurePath + 'MI/'; wbn = ['Reject', 'PVal']; 

	if not os.path.exists(FigurePath2):
		os.makedirs(FigurePath2)

	for wbn2 in wbn:
		PassiveMI = np.zeros(len(Per)); EUMI = np.zeros(len(Per)); EUMI2 = np.zeros(len(Per)); USMI = np.zeros(len(Per))

		if args.DataType=='MNIST' or args.DataType=='MNISTNull':
			PassiveCls = 'NN'
		elif args.DataType=='ADNI':
			PassiveCls = 'logistic'
		elif args.DataType=='Syn':
			PassiveCls = 'logistic'
		Path1 = StatsPath + 'FoundCutEdges(ClsUncertaintyQuery)/Passive/RunDistri/' # assume passive learning results are stored in the SVC folder
		Path2 = StatsPath + 'FoundCutEdges(ClsUncertaintyQuery)/Certainty/%s/RunDistri/InitTrainingSize%d/'%(args.cls, args.InitSize)
		Path3 = StatsPath + 'FoundCutEdges(ClsUncertaintyQuery)/Uncertainty/%s/RunDistri/InitTrainingSize%d/'%(args.cls, args.InitSize)
		Path4 = StatsPath + 'FoundCutEdges(ClsUncertaintyQuery)/Bimodal/%s/RunDistri/InitTrainingSize%d/'%(args.cls, args.InitSize)
	
		pwb=load_workbook(Path1 + 'ClsUncertaintyQuery.xlsx', data_only=True);pws=pwb[wbn2];
		ewb=load_workbook(Path2 + 'ClsUncertaintyQuery.xlsx', data_only=True);ews=ewb[wbn2];
		uwb=load_workbook(Path3 + 'ClsUncertaintyQuery.xlsx', data_only=True);uws=uwb[wbn2];
		ewb2=load_workbook(Path4 + 'ClsUncertaintyQuery.xlsx', data_only=True);ews2=ewb2[wbn2];
		Alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
		# print method name and data type
		print('%s, %s'%(m, args.DataType))
		for i in range(len(Per)):
			for t in range(args.Trial):
				print(t)
				PassiveMI[i]+= pws['%s%d'%(Alphabet[i], t + 1)].value
				EUMI[i]+=ews['%s%d'%(Alphabet[i], t + 1)].value
				USMI[i]+=uws['%s%d'%(Alphabet[i], t + 1)].value
				EUMI2[i]+=ews2['%s%d'%(Alphabet[i], t + 1)].value
		if wbn2 == 'PVal':
			PassiveMI /=args.Trial; EUMI /=args.Trial;USMI /=args.Trial;EUMI2 /=args.Trial
		else:
			ErrorType = 'TypeI'
			if args.DataType == 'Syn' and not (args.Sep == 0 and args.Del == 0):
				ErrorType = 'TypeII'
			elif args.DataType == 'MNIST' or args.DataType == 'MNIST2' or args.DataType == 'ADNI':
				ErrorType = 'TypeII'
			if ErrorType == 'TypeI':
				PassiveMI /=args.Trial; EUMI /=args.Trial;USMI /=args.Trial;EUMI2 /=args.Trial
			elif ErrorType == 'TypeII':
				PassiveMI = args.Trial - PassiveMI; EUMI = args.Trial - EUMI;
				USMI = args.Trial - USMI; EUMI2 = args.Trial - EUMI2;
				PassiveMI /=args.Trial; EUMI /=args.Trial;USMI /=args.Trial;EUMI2 /=args.Trial
		Fig = plt.figure(); ax = Fig.gca()
		ax.plot(Per, PassiveMI, marker = 'o', c='g', markersize = 8, label='Passive', linewidth = 3); 
		ax.plot(Per, USMI, marker = 'o', c='b',markersize = 8, label='Uncertainty', linewidth = 3); 
		ax.plot(Per, EUMI, marker = 'o', c='r', markersize = 8, label='Certainty', linewidth = 3); 
		ax.plot(Per, EUMI2, marker = 'o', c='c', markersize = 8, label='Bimodal(Proposed)', linewidth = 3);
		if wbn2 == 'PVal':
			ax = SetPltProp(ax, xn='Proportion of queries ', yn=wbn2, pos='upper right')
		else:
			if ErrorType == 'TypeI':
				ax.plot(Per, np.ones(len(Per)) * args.alpha, 'o--', c='m', markersize = 8, label='\u03B1=%.2f'%args.alpha, linewidth = 3);
			ax = SetPltProp(ax, xn='Proportion of queries ', yn=ErrorType, pos='upper right')
		Fig.savefig(FigurePath2 + '%s%s%sInitSize%d.png'%(m, wbn2,args.cls, args.InitSize), bbox_inches='tight')
		plt.close('all')

def PlotDimensionTrend(args, FigurePath):
	Per = np.arange(args.Interval, args.Per + args.Interval, args.Interval); 
	FigurePath2 = FigurePath + 'Dimension/'; wbn = ['Reject', 'PVal']
	
	# QueryMethods = ['Passive', 'EnhanceUncertainty', 'uncertainty_sampling', 'EnhanceUncertainty2']
	QueryMethods = ['Passive', 'Bimodal']
	# Label = ['Passive', 'Unimodal', 'Uncertainty', 'Bimodal(Proposed)']; 
	Label = ['Passive', 'Bimodal(Proposed)'];
	# Color = ['g', 'r', 'b', 'c']
	Color = ['g', 'c']
	Dimension = [2, 4, 6, 8, 10, 12, 14, 16, 18]; Alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
	if not os.path.exists(FigurePath2):
		os.makedirs(FigurePath2)
	
	for wbn2 in wbn:
		for i in range(len(Per)):
			Fig = plt.figure(); ax = Fig.gca()
			for count1, q in enumerate(QueryMethods):
				Results = np.zeros(len(Dimension)); Pval= np.zeros((args.Trial, len(Dimension)))		
				for count2, d in enumerate(Dimension):
					if q != 'Passive':
						Path = os.getcwd() + '/Stats/%s/D%dSep%.2fDel%.2fSize%d/FR/Alpha%.2f/FoundCutEdges(ClsUncertaintyQuery)/%s/%s/RunDistri/InitTrainingSize%d/'%(args.DataType, d, args.Sep, args.Del, args.S, args.alpha, q, args.cls, args.InitSize)
					else:
						Path = os.getcwd() + '/Stats/%s/D%dSep%.2fDel%.2fSize%d/FR/Alpha%.2f/FoundCutEdges(ClsUncertaintyQuery)/%s/RunDistri/'%(args.DataType, d, args.Sep, args.Del, args.S, args.alpha, q)
					wb = load_workbook(Path + 'ClsUncertaintyQuery.xlsx', data_only=True);ws = wb[wbn2]
					for t in range(args.Trial):
						# print(q, d, t)
						Results[count2]+= ws['%s%d'%(Alphabet[i], t + 1)].value
						# Pval[t, count2] = ws['%s%d'%(Alphabet[i], t + 1)].value
					if wbn2 == 'PVal':
						Results[count2] /=args.Trial
					else:
						Results[count2] = (args.Trial - Results[count2])/args.Trial

				ax.plot(Dimension, Results, marker = 'o', c=Color[count1], markersize = 8, label=Label[count1], linewidth = 3);
				# if wbn2!= 'Reject':
				# 	Std = np.std(Pval, 0)
				# 	pdb.set_trace()
				# 	ax.fill_between(Dimension, Results + Std, Results - Std, color=Color[count1], alpha = 0.55); 
			if wbn2 == 'Reject':
				ax = SetPltProp(ax, xn='Dimension', yn='TypeII Error', pos='upper left')
			else:
				ax = SetPltProp(ax, xn='Dimension', yn=wbn2, pos='upper left')
			Fig.savefig(FigurePath2 + '%s%.2f.png'%(wbn2, Per[i]), bbox_inches='tight')
			plt.close('all')			 

def PlotFDivergence2(args, Data, QueryPath):
	
	if args.qs != 'Passive':
		MIPath = QueryPath + 'QueryIndex(ClsUncertaintyQuery)/%s/%s/MI/InitTrainingSize%d/'%(args.qs, args.cls, args.InitSize)
		QueryIndexPath = QueryPath + 'QueryIndex(ClsUncertaintyQuery)/%s/%s/RunDistri/InitTrainingSize%d/'%(args.qs, args.cls, args.InitSize)
	else:
		MIPath = QueryPath + 'QueryIndex(ClsUncertaintyQuery)/Passive/MI/'
		QueryIndexPath = QueryPath + 'QueryIndex(ClsUncertaintyQuery)/Passive/RunDistri/'
	Per = np.arange(args.Interval, args.Per + args.Interval, args.Interval); 
	X = Data[:, :-1]; Y = Data[:, -1]; Alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
	
	if not os.path.exists(MIPath):
		os.makedirs(MIPath)

	if os.path.isfile(MIPath + 'MI.xlsx'):
		wb=load_workbook(MIPath + 'MI.xlsx');
		if  not 'FDivergence' in wb.sheetnames or not 'Cutedge' in wb.sheetnames:
			ws1=wb.create_sheet("FDivergence")
			ws2=wb.create_sheet("Cutedge")
		else:
			ws1=wb["FDivergence"]
			ws2=wb["Cutedge"]
	else:
		wb = Workbook(); ws1 = wb.create_sheet("FDivergence"); ws2 = wb.create_sheet("Cutedge"); 

	with open(QueryIndexPath + 'QueryIndex%d.txt'%args.Trial, 'rb') as fp: QueryIndex = pickle.load(fp)
	QueryIndex=np.int64(QueryIndex)
	for i, p in enumerate(Per):
		
		SubX = X[np.sort(QueryIndex[:int(len(QueryIndex) * p)])]; SubY = Y[np.sort(QueryIndex[:int(len(QueryIndex) * p)])];
		I, cutedgenum = FDStats2(args, SubX[SubY==0], SubX[SubY==1]); 	
		ws1.cell(row=args.Trial, column=i+1, value=I)
		ws1['%s%d'%(Alphabet[i], args.Trial + 1)] ='=AVERAGE(%s%d:%s%d)'%(Alphabet[i], 1, Alphabet[i], args.Trial)	

		ws2.cell(row=args.Trial, column=i+1, value=cutedgenum)
		ws2['%s%d'%(Alphabet[i], args.Trial + 1)] ='=AVERAGE(%s%d:%s%d)'%(Alphabet[i], 1, Alphabet[i], args.Trial)	

	wb.save(MIPath + 'MI.xlsx')		


def PlotCI(args, StatsPath, FigurePath):
	if args.qs != 'Passive':
		Path = StatsPath + 'FoundCutEdges(ClsUncertaintyQuery)/%s/%s/RunDistri/InitTrainingSize%d/'%(args.qs, args.cls, args.InitSize)
	else:
		Path = StatsPath + 'FoundCutEdges(ClsUncertaintyQuery)/%s/RunDistri/'%(args.qs)
	FigurePath2 = FigurePath + 'CI/'
	if not os.path.exists(FigurePath2):
		os.makedirs(FigurePath2)
	Per = np.arange(args.Interval, args.Per + args.Interval, args.Interval); 
	Alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
	SuccessNum = np.zeros(len(Per)); Upp_CI = np.zeros(len(Per)); Low_CI = np.zeros(len(Per))

	for t in range(args.Trial):
		Reject = np.load(Path + 'Reject%d.npy'%(t + 1));
		SuccessNum+=Reject

	for i in range(len(Per)):
		
		Low_CI[i], Upp_CI[i] = proportion_confint(SuccessNum[i], args.Trial, alpha = 0.05, method='wilson')
	TypeIError = SuccessNum/args.Trial
	Fig = plt.figure(); ax = Fig.gca()
	if args.qs == 'Passive':
		# ax.plot(Per, TypeIError, marker = 'o', c='b', markersize = 8, label='Passive', linewidth = 3); 
		# ax.fill_between(Per, Low_CI[i], Upp_CI[i], color='b', alpha = 0.15);
		ax.errorbar(Per, TypeIError, yerr=[TypeIError - Low_CI, Upp_CI- TypeIError], 
			        label='Passive', color="black", capsize=3, linestyle="None",
             		marker="s", markersize=7, mfc="black", mec="black")	
	elif args.qs =='Uncertainty':
		ax.errorbar(Per, TypeIError, yerr=[TypeIError - Low_CI, Upp_CI- TypeIError], 
			        label='Uncertainty', color="black", capsize=3, linestyle="None",
             		marker="s", markersize=7, mfc="black", mec="black")	
	elif args.qs == 'Certainty':
		ax.errorbar(Per, TypeIError, yerr=[TypeIError - Low_CI, Upp_CI- TypeIError], 
			        label='Certainty', color="black", capsize=3, linestyle="None",
             		marker="s", markersize=7, mfc="black", mec="black")	
	elif args.qs == 'Bimodal':
		ax.errorbar(Per, TypeIError, yerr=[TypeIError - Low_CI, Upp_CI- TypeIError], 
			        label='Bimodal', color="black", capsize=3, linestyle="None",
             		marker="s", markersize=7, mfc="black", mec="black")	

	ax.plot(Per, np.ones(len(Per)) * args.alpha, c = 'r', label = "Alpha=%.2f"%args.alpha)
	ax = SetPltProp(ax, xn='Query proportion',yn='Type I error', legend = True);
	Fig.savefig(FigurePath2 + 'TypeIErrCI%s%s.png'%(args.qs, args.cls), bbox_inches='tight')
	plt.close('all')	

def WriteToExcel(StatsPath, StatsName, Reject, PVal, Trial):
	if os.path.isfile(StatsPath + '%s.xlsx'%StatsName):
		wb=load_workbook(StatsPath + '%s.xlsx'%StatsName);
		if not 'Reject' in wb.sheetnames:
			ws1=wb.create_sheet("Reject")
		else:
			ws1=wb["Reject"]
		if not 'PVal' in wb.sheetnames:
			ws2=wb.create_sheet("PVal")
		else:
			ws2=wb["PVal"]
	else:
		wb = Workbook(); ws1 = wb.create_sheet("Reject"); ws2 = wb.create_sheet("PVal")
	Alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
	for i in range(len(Reject)):
		ws1.cell(row=Trial, column=i+1, value=Reject[i])
		ws2.cell(row=Trial, column=i+1, value=PVal[i])
		ws1['%s%d'%(Alphabet[i], Trial + 1)] ='=SUM(%s%d:%s%d)'%(Alphabet[i], 1, Alphabet[i], Trial)
		ws2['%s%d'%(Alphabet[i], Trial + 1)] ='=AVERAGE(%s%d:%s%d)'%(Alphabet[i], 1, Alphabet[i], Trial)
	wb.save(StatsPath + '%s.xlsx'%StatsName)

