import numpy as np 
import pdb
from openpyxl import Workbook
from openpyxl import load_workbook
import os 

def main():
	"""
	R: AV45; V: CDRSB; W: ADAS11; X:ADAS13;  Z: MMSE; AA: RAVLT_immediate 
	"""
	Path = os.path.dirname(os.getcwd()) + '/Stats/ADNI/Data/'; 
	ANDIwb =load_workbook(Path + 'ADNIMERGE.xlsx', data_only=True); ANDIws = ANDIwb['ADNIMERGE']
	
	Data = np.zeros((0, 6))
	for i in range(2, ANDIws.max_row):
		if (ANDIws['R%d'%i].value != None and ANDIws['V%d'%i].value != None and  
			ANDIws['W%d'%i].value != None and ANDIws['X%d'%i].value != None and
			ANDIws['Z%d'%i].value != None and ANDIws['AA%d'%i].value != None):
			Data0 = np.array((ANDIws['V%d'%i].value, ANDIws['W%d'%i].value, 
				              ANDIws['X%d'%i].value, ANDIws['Z%d'%i].value,
				              ANDIws['AA%d'%i].value, ANDIws['R%d'%i].value))
			Data = np.vstack((Data, Data0.reshape((1, 6))))
	
	np.save(Path + 'Data.npy', Data)

if __name__=='__main__':
	main()